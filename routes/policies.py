from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, jsonify, abort, session
from flask_login import login_required, current_user
from models import Policy, Client, Product, User, db, UserRole, EmisionStatus, PaymentStatus, Commission
from forms import PolicyForm
from decorators import admin_or_digitador_required, admin_required, admin_or_digitador_or_agent_required
from werkzeug.utils import secure_filename
import openpyxl
import os
import logging
from decimal import Decimal, InvalidOperation
from dateutil import parser  # Importar el módulo para analizar fechas
from sqlalchemy.exc import OperationalError  # Añadir esta línea para importar OperationalError
import sqlalchemy
import sqlalchemy.exc  # Agregar esta línea
from utils.url_helpers import get_return_url
from datetime import datetime, date
from sqlalchemy import or_
import io

bp = Blueprint('policies', __name__, url_prefix='/policies')

@bp.route('/')
@login_required
@admin_or_digitador_or_agent_required
def list_policies():
    # Hacer la sesión permanente al entrar a esta vista
    session.permanent = True
    
    # Guardar/recuperar parámetros de paginación, filtros y ordenamiento en la sesión
    session_key = 'policies_list_params'
    
    # Si hay parámetros en la solicitud, actualizar la sesión
    if request.args:
        # Limpiar la sesión si se hace una nueva búsqueda (cuando hay parámetros pero no hay page)
        if ('policy_number' in request.args or 'client_name' in request.args or 'product_name' in request.args) and 'page' not in request.args:
            if session_key in session:
                session.pop(session_key)
                
        # Guardar los parámetros actuales en la sesión
        session[session_key] = {
            'policy_number': request.args.get('policy_number', ''),
            'client_id': request.args.get('client_id', ''),
            'client_name': request.args.get('client_name', ''),
            'product_id': request.args.get('product_id', ''),
            'product_name': request.args.get('product_name', ''),
            'agent_id': request.args.get('agent_id', ''),
            'agent_name': request.args.get('agent_name', ''),
            'status': request.args.get('status', ''),
            'payment_status': request.args.get('payment_status', ''),
            'emision_status': request.args.get('emision_status', ''),
            'sort_by': request.args.get('sort_by', 'start_date'),
            'sort_order': request.args.get('sort_order', 'desc'),
            'page': request.args.get('page', 1, type=int),
            'per_page': request.args.get('per_page', 10, type=int)
        }
    
    # Si no hay parámetros pero sí hay sesión guardada, recuperarla para mantener el estado
    elif session_key in session:
        return redirect(url_for('policies.list_policies', **session[session_key]))
    
    # Obtener los parámetros ya sea de la solicitud o de la sesión
    params = session.get(session_key, {})
    policy_number = params.get('policy_number', request.args.get('policy_number', ''))
    client_id = params.get('client_id', request.args.get('client_id', ''))
    client_name = params.get('client_name', request.args.get('client_name', ''))
    product_id = params.get('product_id', request.args.get('product_id', ''))
    product_name = params.get('product_name', request.args.get('product_name', ''))
    agent_id = params.get('agent_id', request.args.get('agent_id', ''))
    agent_name = params.get('agent_name', request.args.get('agent_name', ''))
    status = params.get('status', request.args.get('status', ''))
    payment_status = params.get('payment_status', request.args.get('payment_status', ''))
    emision_status = params.get('emision_status', request.args.get('emision_status', ''))
    
    # Parámetros de ordenamiento
    sort_by = params.get('sort_by', request.args.get('sort_by', 'start_date'))
    sort_order = params.get('sort_order', request.args.get('sort_order', 'desc'))
    
    # Mapeo de nombres de parámetros a atributos de modelo para ordenamiento
    sort_columns = {
        'policy_number': Policy.policy_number,
        'client': Client.name,
        'product': Product.name,
        'premium': Policy.premium,
        'start_date': Policy.start_date,
        'end_date': Policy.end_date,
        'agent': User.name,
        'emision_status': Policy.emision_status,
        'payment_status': Policy.payment_status,
        'solicitation_date': Policy.solicitation_date
    }
    
    # Aplicar joins para filtrado y ordenamiento
    query = Policy.query
    
    # Siempre aplicamos los joins aquí para poder filtrar por relaciones
    query = query.join(Client, Policy.client_id == Client.id)
    query = query.join(Product, Policy.product_id == Product.id)
    query = query.join(User, Policy.agent_id == User.id)
    
    # Aplicar filtros
    if policy_number:
        query = query.filter(Policy.policy_number.ilike(f'%{policy_number}%'))
    
    # Filtrado de cliente (por ID o por nombre)
    if client_id:
        query = query.filter(Policy.client_id == client_id)
    elif client_name:
        query = query.filter(Client.name.ilike(f'%{client_name}%'))
    
    # Filtrado de producto (por ID o por nombre)
    if product_id:
        query = query.filter(Policy.product_id == product_id)
    elif product_name:
        query = query.filter(Product.name.ilike(f'%{product_name}%'))
    
    # Filtrado de agente (por ID o por nombre)
    if agent_id:
        query = query.filter(Policy.agent_id == agent_id)
    elif agent_name:
        query = query.filter(User.name.ilike(f'%{agent_name}%'))
    
    if status:
        query = query.filter(Policy.status == status)
    if payment_status:
        query = query.filter(Policy.payment_status == payment_status)
    if emision_status:
        query = query.filter(Policy.emision_status == emision_status)
    
    # Aplicar ordenamiento
    if sort_by in sort_columns:
        if sort_order == 'desc':
            query = query.order_by(sort_columns[sort_by].desc())
        else:
            query = query.order_by(sort_columns[sort_by].asc())
    else:
        query = query.order_by(Policy.start_date.desc())
    
    page = params.get('page', request.args.get('page', 1, type=int))
    per_page = params.get('per_page', request.args.get('per_page', 10, type=int))
    
    allowed_per_page = [10, 25, 50, 100]
    if per_page not in allowed_per_page:
        per_page = 10
    
    pagination = query.paginate(
        page=page, 
        per_page=per_page,
        error_out=False
    )
    
    return render_template('policies/list.html', 
                          policies=pagination.items, 
                          pagination=pagination,
                          clients=Client.query.all(),
                          products=Product.query.all(),
                          agents=User.query.filter_by(role=UserRole.AGENTE).all(),
                          title="Lista de Pólizas",
                          sort_by=sort_by,
                          sort_order=sort_order,
                          EmisionStatus=EmisionStatus,
                          PaymentStatus=PaymentStatus,
                          client_id=client_id,
                          client_name=client_name,
                          product_id=product_id,
                          product_name=product_name,
                          agent_id=agent_id,
                          agent_name=agent_name)

@bp.route('/create', methods=['GET', 'POST'])
@login_required
@admin_or_digitador_required
def create_policy():
    form = PolicyForm()
    # Cargar las opciones de productos
    products = Product.query.all()
    form.product_id.choices = [(p.id, p.name) for p in products]
    
    if form.validate_on_submit():
        try:
            # Convertir los valores de los select a enumeraciones
            emision_status_enum = EmisionStatus[form.emision_status.data]
            payment_status_enum = PaymentStatus[form.payment_status.data]
            
            # Crear la póliza con los datos ya convertidos
            policy = Policy(
                policy_number=form.policy_number.data,
                start_date=form.start_date.data,
                end_date=form.end_date.data,
                premium=form.premium.data,
                product_id=form.product_id.data,
                client_id=form.client_id.data,
                agent_id=form.agent_id.data,
                emision_status=emision_status_enum,
                payment_status=payment_status_enum,
                solicitation_date=form.solicitation_date.data
            )
            
            if current_user.role == UserRole.AGENTE:
                policy.agent_id = current_user.id
            
            db.session.add(policy)
            db.session.commit()
            flash('Póliza creada exitosamente', 'success')
            return redirect(get_return_url(url_for('policies.list_policies')))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear la póliza: {str(e)}', 'error')
    
    return render_template('policies/create.html', form=form)

@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_or_digitador_required
def edit_policy(id):
    policy = Policy.query.get_or_404(id)
    form = PolicyForm(obj=policy)
    # Establecer el ID de la póliza para la validación
    form.id.data = str(policy.id)
    
    # Cargar las opciones de productos
    products = Product.query.all()
    form.product_id.choices = [(p.id, p.name) for p in products]
    
    # Cargar el cliente y agente actuales si el formulario no ha sido enviado
    if not form.is_submitted():
        if policy.client:
            form.client.data = policy.client.name
            form.client_id.data = policy.client_id
        if policy.agent:
            form.agent.data = policy.agent.name
            form.agent_id.data = policy.agent_id
    
    if form.validate_on_submit():
        try:
            # En lugar de usar populate_obj, actualizar manualmente cada campo
            policy.policy_number = form.policy_number.data
            policy.start_date = form.start_date.data
            policy.end_date = form.end_date.data
            policy.premium = form.premium.data
            policy.client_id = form.client_id.data
            policy.product_id = form.product_id.data
            policy.agent_id = form.agent_id.data
            policy.solicitation_date = form.solicitation_date.data
            
            # Convertir los valores de los select a enumeraciones
            policy.emision_status = EmisionStatus[form.emision_status.data]
            policy.payment_status = PaymentStatus[form.payment_status.data]
            
            db.session.commit()
            flash('Póliza actualizada exitosamente', 'success')
            return redirect(get_return_url(url_for('policies.list_policies')))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar la póliza: {str(e)}', 'error')
    
    return render_template('policies/edit.html', form=form, policy=policy)

@bp.route('/delete/<int:id>')
@login_required
@admin_required
def delete_policy(id):
    policy = Policy.query.get_or_404(id)
    try:
        db.session.delete(policy)
        db.session.commit()
        flash('Póliza eliminada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la póliza: {str(e)}', 'error')
    
    return redirect(get_return_url(url_for('policies.list_policies')))

@bp.route('/bulk_upload', methods=['GET', 'POST'])
@login_required
@admin_or_digitador_required
def bulk_upload_policies():
    if request.method == 'POST':
        # Si estamos en la fase de mapeo de columnas, procesar directamente
        if 'column_mapping' in request.form:
            # Recuperar el archivo temporal de la sesión
            file_path = session.get('temp_file_path')
            if not file_path:
                flash('Error: No se encontró el archivo temporal. Por favor, suba el archivo nuevamente.', 'error')
                return redirect(url_for('policies.bulk_upload_policies', _external=True))
                
            return process_file_with_mapping(None, request.form)
            
        # Primera fase: validar y subir el archivo
        if 'file' not in request.files:
            flash('No se ha seleccionado ningún archivo', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No se ha seleccionado ningún archivo', 'error')
            return redirect(request.url)
            
        if file and file.filename.endswith('.xlsx'):
            # Primera fase: analizar el archivo y mostrar la pantalla de mapeo
            try:
                filename = secure_filename(file.filename)
                file_path = os.path.join('/tmp', filename)
                file.save(file_path)
                
                # Guardar la ruta del archivo en la sesión para usarla en el siguiente paso
                session['temp_file_path'] = file_path
                
                # Leer encabezados del archivo
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                
                # Obtener los encabezados (primera fila)
                headers = []
                for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
                    headers.append(cell if cell else "")
                
                # Obtener los datos de muestra (segunda fila si existe)
                sample_data = []
                second_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
                if second_row:
                    sample_data = list(second_row[0])
                
                # Campos requeridos del sistema con sus tipos de datos
                system_fields = [
                    {"id": "policy_number", "name": "Número de Póliza", "type": "text", "required": True},
                    {"id": "start_date", "name": "Fecha de Inicio", "type": "date", "required": True},
                    {"id": "end_date", "name": "Fecha de Fin", "type": "date", "required": True},
                    {"id": "premium", "name": "Prima", "type": "decimal", "required": True},
                    {"id": "client_document_number", "name": "Número de Documento del Cliente", "type": "text", "required": True},
                    {"id": "product_name", "name": "Nombre del Producto", "type": "text", "required": True},
                    {"id": "agent_document_number", "name": "Número de Documento del Agente", "type": "text", "required": True},
                    {"id": "emision_status", "name": "Estado de Emisión", "type": "enum", "required": True},
                    {"id": "payment_status", "name": "Estado de Pago", "type": "enum", "required": True},
                    {"id": "solicitation_date", "name": "Fecha de Solicitud", "type": "date", "required": False},
                ]
                
                return render_template(
                    'policies/column_mapping.html',
                    headers=headers, 
                    sample_data=sample_data,
                    system_fields=system_fields,
                    form_action=url_for('policies.bulk_upload_policies', _external=True)  # URL absoluta para el formulario
                )
                
            except Exception as e:
                flash(f'Error al analizar el archivo: {str(e)}', 'error')
                return redirect(url_for('policies.bulk_upload_policies', _external=True))
        else:
            flash('Formato de archivo inválido. Por favor, suba un archivo XLSX.', 'error')
            
    return render_template('policies/bulk_upload.html')

def process_file_with_mapping(file, form_data):
    """Procesa el archivo Excel utilizando el mapeo de columnas proporcionado por el usuario"""
    
    # Recuperar la ruta del archivo temporal
    file_path = session.get('temp_file_path')
    if not file_path:
        flash('Error: No se encontró el archivo temporal. Por favor, suba el archivo nuevamente.', 'error')
        return redirect(url_for('policies.bulk_upload_policies', _external=True))
    
    # Verificar si el archivo existe
    if not os.path.exists(file_path):
        flash('Error: El archivo temporal no existe. Por favor, suba el archivo nuevamente.', 'error')
        if 'temp_file_path' in session:
            session.pop('temp_file_path')
        return redirect(url_for('policies.bulk_upload_policies', _external=True))
    
    # Obtener el mapeo de columnas del formulario
    column_mapping = {}
    for field_id in form_data:
        if field_id.startswith('field_'):
            system_field = field_id.replace('field_', '')
            excel_column_index = int(form_data[field_id])
            if excel_column_index >= 0:  # -1 significa "No mapear esta columna"
                column_mapping[system_field] = excel_column_index
    
    all_error_details = []
    updated_count = 0
    created_count = 0
    error_count = 0
        
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Procesar cada fila, empezando desde la segunda (la primera son los encabezados)
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            # Iniciar una nueva transacción para cada fila
            try:
                # Mapear los valores de la fila según la configuración del usuario
                policy_data = {}
                
                # Para cada campo del sistema, obtener el valor de la columna correspondiente
                for system_field, excel_index in column_mapping.items():
                    if excel_index < len(row):
                        policy_data[system_field] = row[excel_index]
                    else:
                        policy_data[system_field] = None
                
                # Log para depuración
                logging.info(f"Procesando fila {idx} con datos mapeados: {policy_data}")
                
                # Extraer valores específicos para validación
                policy_number = policy_data.get('policy_number')
                client_document_number = policy_data.get('client_document_number')
                product_name = policy_data.get('product_name')
                agent_document_number = policy_data.get('agent_document_number')
                
                # Verificar si los datos obligatorios están presentes
                if not all([policy_number, client_document_number, product_name, agent_document_number]):
                    raise ValueError("Faltan datos obligatorios (número de póliza, documento del cliente, producto o agente)")
                
                # Convertir a string si no lo es
                if client_document_number is not None and not isinstance(client_document_number, str):
                    client_document_number = str(int(client_document_number))
                
                if agent_document_number is not None and not isinstance(agent_document_number, str):
                    agent_document_number = str(int(agent_document_number))
                
                # Buscar cliente, producto y agente
                client = Client.query.filter_by(document_number=client_document_number).first()
                if not client:
                    raise ValueError(f"Cliente con documento {client_document_number} no encontrado")
                
                product = Product.query.filter_by(name=product_name).first()
                if not product:
                    raise ValueError(f"Producto '{product_name}' no encontrado")
                
                agent = User.query.filter_by(document_number=agent_document_number, role=UserRole.AGENTE).first()
                if not agent:
                    raise ValueError(f"Agente con documento {agent_document_number} no encontrado")
                
                # Procesar valores específicos
                premium = policy_data.get('premium')
                start_date = policy_data.get('start_date')
                end_date = policy_data.get('end_date')
                emision_status = policy_data.get('emision_status')
                payment_status = policy_data.get('payment_status')
                solicitation_date = policy_data.get('solicitation_date')
                
                try:
                    premium_decimal = Decimal(str(premium))
                except InvalidOperation:
                    raise ValueError(f"Valor de prima inválido: {premium}")
                
                # Procesar fechas
                try:
                    # Verificar si las fechas ya son objetos date o datetime
                    if isinstance(start_date, (datetime, date)):
                        start_date_parsed = start_date.date() if isinstance(start_date, datetime) else start_date
                    else:
                        start_date_parsed = parser.parse(str(start_date)).date()
                        
                    if isinstance(end_date, (datetime, date)):
                        end_date_parsed = end_date.date() if isinstance(end_date, datetime) else end_date
                    else:
                        end_date_parsed = parser.parse(str(end_date)).date()
                except Exception as e:
                    raise ValueError(f"Error al procesar las fechas: {str(e)}")
                
                # Procesar fecha de solicitud si existe
                solicitation_date_parsed = None
                if solicitation_date:
                    try:
                        if isinstance(solicitation_date, (datetime, date)):
                            solicitation_date_parsed = solicitation_date.date() if isinstance(solicitation_date, datetime) else solicitation_date
                        else:
                            solicitation_date_parsed = parser.parse(str(solicitation_date)).date()
                    except:
                        logging.warning(f"Fecha de solicitud inválida en fila {idx}: {solicitation_date}")
                
                # Procesar estado de emisión
                try:
                    # Mapeo para valores que existen en el modelo pero no en la base de datos
                    emision_status_map = {
                        'REQ': 'PENDIENTE',            # Mapear REQ a PENDIENTE
                        'REENVIADA': 'PENDIENTE',      # Mapear REENVIADA a PENDIENTE
                        'req': 'PENDIENTE',            # Versión en minúsculas
                        'reenviada': 'PENDIENTE',      # Versión en minúsculas
                        'Req': 'PENDIENTE',            # Versión mixta
                        'Reenviada': 'PENDIENTE',      # Versión mixta
                    }
                    
                    # Normalizar a mayúsculas
                    emision_status_upper = emision_status.upper()
                    
                    # Si necesita mapeo, hacer la conversión
                    if emision_status_upper in emision_status_map:
                        emision_status_upper = emision_status_map[emision_status_upper]
                        logging.info(f"Estado de emisión mapeado: '{emision_status}' -> {emision_status_upper}")
                    
                    # Ahora buscar el enum por nombre
                    try:
                        emision_status_enum = EmisionStatus[emision_status_upper]
                        logging.info(f"Estado de emisión encontrado: {emision_status_enum.name}")
                    except KeyError:
                        # Si no se encuentra, usar PENDIENTE por defecto
                        logging.warning(f"Estado de emisión no encontrado: '{emision_status}', usando PENDIENTE")
                        emision_status_enum = EmisionStatus.PENDIENTE
                        
                except Exception as e:
                    logging.error(f"Error procesando estado de emisión: {str(e)}")
                    # En caso de error, usar PENDIENTE como fallback
                    emision_status_enum = EmisionStatus.PENDIENTE
                    logging.warning(f"Usando valor fallback PENDIENTE para '{emision_status}'")
                
                # Procesar estado de pago
                try:
                    payment_status_enum = PaymentStatus[payment_status.upper()]
                except KeyError:
                    # Intentar buscar por valor
                    try:
                        found = False
                        for status in PaymentStatus:
                            if status.value.upper() == payment_status.upper():
                                payment_status_enum = status
                                logging.info(f"Estado de pago mapeado: '{payment_status}' -> {status.name}")
                                found = True
                                break
                        if not found:
                            valid_states = [f"{p.name} ({p.value})" for p in PaymentStatus]
                            raise ValueError(f"Estado de pago no válido: '{payment_status}'. Estados válidos: {', '.join(valid_states)}")
                    except Exception as e:
                        raise ValueError(f"Estado de pago no válido: '{payment_status}'. Error: {str(e)}")
                
                # Comprobar si la póliza existe - usar una nueva transacción interna
                try:
                    # Asegurarse de que policy_number sea un string
                    if not isinstance(policy_number, str):
                        policy_number = str(policy_number)
                        
                    existing_policy = Policy.query.filter_by(policy_number=policy_number).first()
                    if existing_policy:
                        existing_policy.start_date = start_date_parsed
                        existing_policy.end_date = end_date_parsed
                        existing_policy.premium = float(premium_decimal)
                        existing_policy.client_id = client.id
                        existing_policy.product_id = product.id
                        existing_policy.agent_id = agent.id
                        existing_policy.emision_status = emision_status_enum
                        existing_policy.payment_status = payment_status_enum
                        existing_policy.solicitation_date = solicitation_date_parsed
                        db.session.commit()
                        updated_count += 1
                        logging.info(f'Póliza {policy_number} actualizada exitosamente.')
                        
                        # Calcular comisión si el estado de pago es PAGADO
                        if payment_status_enum == PaymentStatus.PAGADO:
                            try:
                                commission = agent.calculate_commission(existing_policy)
                                if commission:
                                    logging.info(f'Comisión calculada para la póliza {policy_number}: {commission.amount}')
                            except Exception as comm_error:
                                logging.error(f'Error al calcular comisión para póliza {policy_number}: {str(comm_error)}')
                    else:
                        policy = Policy(
                            policy_number=policy_number,
                            start_date=start_date_parsed,
                            end_date=end_date_parsed,
                            premium=float(premium_decimal),
                            client_id=client.id,
                            product_id=product.id,
                            agent_id=agent.id,
                            emision_status=emision_status_enum,
                            payment_status=payment_status_enum,
                            solicitation_date=solicitation_date_parsed
                        )
                        db.session.add(policy)
                        db.session.commit()
                        created_count += 1
                        logging.info(f'Póliza {policy_number} creada exitosamente.')
                        
                        # Calcular comisión si el estado de pago es PAGADO
                        if payment_status_enum == PaymentStatus.PAGADO:
                            try:
                                commission = agent.calculate_commission(policy)
                                if commission:
                                    logging.info(f'Comisión calculada para la póliza {policy_number}: {commission.amount}')
                            except Exception as comm_error:
                                logging.error(f'Error al calcular comisión para póliza {policy_number}: {str(comm_error)}')
                except Exception as e:
                    # Revertir transacción y registrar error
                    db.session.rollback()
                    logging.error(f"Error al crear/actualizar póliza {policy_number}: {str(e)}")
                    raise ValueError(f"Error al crear/actualizar póliza {policy_number}: {str(e)}")
            
            except Exception as e:
                # Registrar error y continuar con la siguiente fila
                error_count += 1
                error_detail = {
                    'row': idx,
                    'message': str(e),
                    'data': str(row) if row else "N/A"
                }
                all_error_details.append(error_detail)
                logging.error(f"Error procesando fila {idx}: {str(e)}")
                db.session.rollback()  # Asegurar que el error no afecte la siguiente transacción
        
        # Enviar el mensaje flash después del commit
        if created_count > 0 or updated_count > 0 or error_count > 0:
            # Limpiar archivo temporal
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as cleanup_error:
                logging.error(f"Error al limpiar archivo temporal: {str(cleanup_error)}")
            
            # Eliminar el archivo temporal de la sesión
            if 'temp_file_path' in session:
                session.pop('temp_file_path')
            
            # Mostrar resumen de la operación
            flash(f'Procesamiento completado: {created_count} pólizas creadas, {updated_count} actualizadas, {error_count} errores.', 'info' if error_count == 0 else 'error')
            
            # Preparar resumen para la plantilla
            summary = {
                'created': created_count,
                'updated': updated_count,
                'errors': error_count,
                'error_details': all_error_details[:10]  # Limitar a los primeros 10 errores para no sobrecargar la página
            }
            
            return render_template('policies/bulk_upload.html', summary=summary)
        else:
            flash('No se procesaron registros. Verifique el formato del archivo.', 'warning')
            return redirect(url_for('policies.bulk_upload_policies'))
        
    except Exception as e:
        # Error general en el procesamiento del archivo
        flash(f'Error general al procesar el archivo: {str(e)}', 'error')
        logging.error(f"Error general en la carga masiva: {str(e)}")
        # Limpiar archivo temporal
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception as cleanup_error:
            logging.error(f"Error al limpiar archivo temporal: {str(cleanup_error)}")
        return redirect(url_for('policies.bulk_upload_policies', _external=True))

@bp.route('/download_sample')
@login_required
@admin_or_digitador_required
def download_policy_sample():
    return send_file('static/samples/polizas_ejemplo.xlsx', as_attachment=True)

@bp.route('/export_config')
@login_required
@admin_or_digitador_required
def export_config():
    """
    Muestra el formulario para configurar la exportación de pólizas
    permitiendo al usuario seleccionar qué columnas incluir
    """
    return render_template('policies/export_config.html')

@bp.route('/export_policies')
@login_required
@admin_or_digitador_required
def export_policies_for_update():
    """
    Exporta las pólizas filtradas a un archivo Excel con las columnas seleccionadas por el usuario
    """
    # Obtener parámetros de filtro de la solicitud
    policy_number = request.args.get('policy_number', '')
    client_id = request.args.get('client_id', '')
    client_name = request.args.get('client_name', '')
    product_id = request.args.get('product_id', '')
    product_name = request.args.get('product_name', '')
    agent_id = request.args.get('agent_id', '')
    agent_name = request.args.get('agent_name', '')
    emision_status = request.args.get('emision_status', '')
    payment_status = request.args.get('payment_status', '')
    start_date_from = request.args.get('start_date_from', '')
    start_date_to = request.args.get('start_date_to', '')
    end_date_from = request.args.get('end_date_from', '')
    end_date_to = request.args.get('end_date_to', '')
    solicitation_date_from = request.args.get('solicitation_date_from', '')
    solicitation_date_to = request.args.get('solicitation_date_to', '')
    
    # Obtener columnas seleccionadas
    selected_columns = request.args.getlist('columns')
    
    # Si no hay columnas seleccionadas, redirigir a la página de configuración
    if not selected_columns:
        flash('Por favor, seleccione al menos una columna para exportar', 'warning')
        return redirect(url_for('policies.export_config', **request.args))
    
    # Construir la consulta
    query = Policy.query
    
    # Aplicar los mismos filtros que en list_policies
    query = query.join(Client, Policy.client_id == Client.id)
    query = query.join(Product, Policy.product_id == Product.id)
    query = query.join(User, Policy.agent_id == User.id)
    
    if policy_number:
        query = query.filter(Policy.policy_number.ilike(f'%{policy_number}%'))
    
    if client_id:
        query = query.filter(Policy.client_id == client_id)
    elif client_name:
        query = query.filter(Client.name.ilike(f'%{client_name}%'))
    
    if product_id:
        query = query.filter(Policy.product_id == product_id)
    elif product_name:
        query = query.filter(Product.name.ilike(f'%{product_name}%'))
    
    if agent_id:
        query = query.filter(Policy.agent_id == agent_id)
    elif agent_name:
        query = query.filter(User.name.ilike(f'%{agent_name}%'))
    
    if emision_status:
        query = query.filter(Policy.emision_status == emision_status)
    
    if payment_status:
        query = query.filter(Policy.payment_status == payment_status)
    
    # Filtrar por fechas
    if start_date_from:
        try:
            start_from = parser.parse(start_date_from).date()
            query = query.filter(Policy.start_date >= start_from)
        except:
            pass
    
    if start_date_to:
        try:
            start_to = parser.parse(start_date_to).date()
            query = query.filter(Policy.start_date <= start_to)
        except:
            pass
    
    if end_date_from:
        try:
            end_from = parser.parse(end_date_from).date()
            query = query.filter(Policy.end_date >= end_from)
        except:
            pass
    
    if end_date_to:
        try:
            end_to = parser.parse(end_date_to).date()
            query = query.filter(Policy.end_date <= end_to)
        except:
            pass
    
    if solicitation_date_from:
        try:
            sol_from = parser.parse(solicitation_date_from).date()
            query = query.filter(Policy.solicitation_date >= sol_from)
        except:
            pass
    
    if solicitation_date_to:
        try:
            sol_to = parser.parse(solicitation_date_to).date()
            query = query.filter(Policy.solicitation_date <= sol_to)
        except:
            pass
    
    # Obtener las pólizas
    policies = query.order_by(Policy.policy_number).all()
    
    if not policies:
        flash('No se encontraron pólizas con los filtros seleccionados.', 'warning')
        return redirect(url_for('policies.list_policies'))
    
    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pólizas"
    
    # Mapeo de valores de columnas seleccionadas a encabezados y funciones de obtención de datos
    column_mappings = {
        # Información de Póliza
        'policy_number': {
            'header': "Número de Póliza",
            'getter': lambda p: p.policy_number
        },
        'start_date': {
            'header': "Fecha de Inicio",
            'getter': lambda p: p.start_date
        },
        'end_date': {
            'header': "Fecha de Fin",
            'getter': lambda p: p.end_date
        },
        'premium': {
            'header': "Prima",
            'getter': lambda p: float(p.premium)
        },
        'solicitation_date': {
            'header': "Fecha de Solicitud",
            'getter': lambda p: p.solicitation_date
        },
        
        # Información del Cliente
        'client_name': {
            'header': "Nombre del Cliente",
            'getter': lambda p: p.client.name if p.client else ''
        },
        'client_document_number': {
            'header': "Número de Documento del Cliente",
            'getter': lambda p: p.client.document_number if p.client else ''
        },
        'client_email': {
            'header': "Email del Cliente",
            'getter': lambda p: p.client.email if p.client else ''
        },
        'client_phone': {
            'header': "Teléfono del Cliente",
            'getter': lambda p: p.client.phone if p.client else ''
        },
        'client_address': {
            'header': "Dirección del Cliente",
            'getter': lambda p: p.client.address if p.client else ''
        },
        
        # Información del Producto
        'product_name': {
            'header': "Nombre del Producto",
            'getter': lambda p: p.product.name if p.product else ''
        },
        'product_description': {
            'header': "Descripción del Producto",
            'getter': lambda p: p.product.description if p.product else ''
        },
        'product_aseguradora': {
            'header': "Aseguradora",
            'getter': lambda p: p.product.aseguradora if p.product else ''
        },
        
        # Información del Agente
        'agent_name': {
            'header': "Nombre del Agente",
            'getter': lambda p: p.agent.name if p.agent else ''
        },
        'agent_document_number': {
            'header': "Número de Documento del Agente",
            'getter': lambda p: p.agent.document_number if p.agent else ''
        },
        'agent_email': {
            'header': "Email del Agente",
            'getter': lambda p: p.agent.email if p.agent else ''
        },
        'agent_phone': {
            'header': "Teléfono del Agente",
            'getter': lambda p: p.agent.phone if p.agent else ''
        },
        
        # Estados
        'emision_status': {
            'header': "Estado de Emisión",
            'getter': lambda p: p.emision_status.name if p.emision_status else ''
        },
        'payment_status': {
            'header': "Estado de Pago",
            'getter': lambda p: p.payment_status.name if p.payment_status else ''
        }
    }
    
    # Filtrar solo las columnas seleccionadas y asegurar que existan en el mapeo
    headers = []
    getters = []
    for col in selected_columns:
        if col in column_mappings:
            headers.append(column_mappings[col]['header'])
            getters.append(column_mappings[col]['getter'])
    
    # Si no hay columnas válidas, usar valores por defecto
    if not headers:
        headers = ["Número de Póliza", "Fecha de Inicio", "Fecha de Fin", "Prima"]
        getters = [
            lambda p: p.policy_number,
            lambda p: p.start_date,
            lambda p: p.end_date,
            lambda p: float(p.premium)
        ]
    
    # Escribir los encabezados
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Escribir los datos de las pólizas
    for row_idx, policy in enumerate(policies, 2):
        for col_idx, getter in enumerate(getters, 1):
            ws.cell(row=row_idx, column=col_idx, value=getter(policy))
    
    # Guardar el archivo en un búfer en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generar un nombre de archivo con timestamp
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"polizas_export_{timestamp}.xlsx"
    
    # Enviar el archivo
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/search_clients')
@login_required
@admin_or_digitador_required
def search_clients():
    query = request.args.get('query', '')
    clients = Client.query.filter(Client.name.ilike(f'%{query}%')).limit(10).all()
    return jsonify([{'id': c.id, 'name': c.name} for c in clients])

@bp.route('/search_agents')
@login_required
@admin_or_digitador_required
def search_agents():
    query = request.args.get('query', '')
    agents = User.query.filter(
        User.role == UserRole.AGENTE,
        User.name.ilike(f'%{query}%')
    ).limit(10).all()
    return jsonify([{
        'id': a.id, 
        'name': f"{a.name} ({a.document_number})"
    } for a in agents])

@bp.route('/<int:id>')
@login_required
@admin_or_digitador_or_agent_required
def policy_detail(id):
    """Vista detallada de una póliza."""
    policy = Policy.query.get_or_404(id)
    
    # Verificar si el usuario actual es un agente y si la póliza le pertenece
    if current_user.role == UserRole.AGENTE and policy.agent_id != current_user.id:
        abort(403)  # Forbidden
    
    # Calcular comisiones
    commissions = Commission.query.filter_by(policy_id=policy.id).all()
    total_commission = sum(float(commission.amount) for commission in commissions)
    
    return render_template(
        'policies/detail.html',
        policy=policy,
        commissions=commissions,
        total_commission=total_commission
    )

@bp.route('/recalculate_commission/<int:id>')
@login_required
@admin_required
def recalculate_commission(id):
    try:
        policy = Policy.query.get_or_404(id)
        
        # Eliminar comisiones existentes
        Commission.query.filter_by(policy_id=policy.id).delete()
        
        # Recalcular comisiones
        agent = User.query.get(policy.agent_id)
        commission = agent.calculate_commission(policy)
        
        if commission:
            db.session.commit()
            flash('Comisiones recalculadas exitosamente.', 'success')
        else:
            flash('No se pudieron calcular las comisiones.', 'error')
            
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error al recalcular comisiones: {str(e)}")
        flash(f'Error al recalcular comisiones: {str(e)}', 'error')
    
    return redirect(get_return_url(url_for('policies.policy_detail', id=id)))

@bp.route('/search_products')
@login_required
@admin_or_digitador_or_agent_required
def search_products():
    query = request.args.get('query', '')
    products = Product.query.filter(Product.name.ilike(f'%{query}%')).limit(10).all()
    return jsonify([{'id': p.id, 'name': p.name} for p in products])

